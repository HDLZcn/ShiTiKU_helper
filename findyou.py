import openpyxl
import os
import pyperclip
import time
print(os.getcwd())

#============载入阶段===================def LoadXlsx():
#载入一个xlsx文件
wb=openpyxl.load_workbook("findme.xlsx")
#获取xlsx文件中活动表，并打印其表名
activeSheet = str(wb.active.title)
aSheet = wb.active
print('activesheet is '+activeSheet+'.\n')
#建立一个字典data
loadedDataDict={0:"题目"}
#遍历这个活动表的A列，讲A列的内容以{1：“题目”}的方式存入字典中
maxRow=wb[activeSheet].max_row+42
for i in range (1,maxRow) :
    sheetValue=  aSheet.cell(row= i , column=1).value
    loadedDataDict[ i ] = sheetValue
#ivd = {v: k for k, v in loadedDataDict.items()}
#print("loadedDataDict[50]:",loadedDataDict[50])
#============载入完毕===================
#定义一个函数，输入key输出一个列表M。具体为A（key）行的每个内容：题目、答案、选项12345
    #+++写完才发现这个用列表+循环一下就能搞好，我在这傻乎乎print啥呢这是？
def findKey (targetKey):
    print("++++++++++++++++++++++++++++++++++++++++++++")
    print("+Q:"+loadedDataDict[targetKey]+"+"+".\n")
    print("+=============================="+"+")
    print("+"+str(aSheet.cell(row= 1 , column=2).value)+":"+str(aSheet.cell(row= targetKey , column=2).value)+".\n")
    print("+"+str(aSheet.cell(row= 1 , column=3).value)+":"+str(aSheet.cell(row= targetKey , column=3).value)+".\n")
    print("+"+str(aSheet.cell(row= 1 , column=4).value)+":"+str(aSheet.cell(row= targetKey , column=4).value)+".\n")
    print("+"+str(aSheet.cell(row= 1 , column=5).value)+":"+str(aSheet.cell(row= targetKey , column=5).value)+".\n")
    print("+"+str(aSheet.cell(row= 1 , column=6).value)+":"+str(aSheet.cell(row= targetKey , column=6).value)+".\n")
    print("+"+str(aSheet.cell(row= 1 , column=7).value)+":"+str(aSheet.cell(row= targetKey , column=7).value)+".\n")
    print("+"+str(aSheet.cell(row= 1 , column=8).value)+":"+str(aSheet.cell(row= targetKey , column=8).value)+".\n")
    print("++++++++++++++++++++++++++++++++++++++++++++")

#===========剪贴板嗅探开始===============
def snifferClip():
#嗅探剪贴板的最新内容
    temp = 1
    tempNew = pyperclip.paste()
    if tempNew == "注销":
        return -1
#存入temp变量中，检查是否有变化
    elif temp == tempNew:
        return 0
    else:
        temp != tempNew
        return 1

#===========剪贴板嗅探结束===============
#===========模糊搜索开始===============
#temp文件作为模糊词输入函数match
def wantFind():
    wantFind = str(pyperclip.paste())
    x=0
    y=0
    maxx=wb[activeSheet].max_row+42
    for x,y in loadedDataDict.items():
        try:
            if wantFind in y:
                        
                return x
            else:
                x=x+1
                
        except Exception as TypeError:
            print("====啥也没查到，这个好像不在题库中！=========\n====重启程序查询下一题======================\n====或者复制时跳过标点符号如。，（）再试=====")
            return 1

                
#模糊搜索在data中进行搜索value
#返回key
#===========模糊搜索结束===============
#隔0.5秒刷新一次
#主函数：
#打印列表M LoadXlsx()
#等待下一次传入
def findYouMain ():
    while snifferClip() > 0:
        if snifferClip() == 0:
            time.sleep(0.5)
            print('我找到了！快看看吧！')
        else:
            print('已经嗅探到你复制的东西了，正在查询全部【'+activeSheet+'】……\n')
            findKey(targetKey = int(wantFind()))
            time.sleep(2)
           
print ("=====================详细用法见readme.txt================\n欢迎使用试题库刷题小工具\n作者：hdlz\n本工具只用于学习交流使用，如使用此工具进行违法乱纪行为，本作者不承担任何连带责任\n4秒后开始刷题\n当你想安全退出的时候，请复制注销两个字\n\n\n=====================详细用法见readme.txt================\n祝你玩的开心！")
time.sleep(5)
findYouMain ()
